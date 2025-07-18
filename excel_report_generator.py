import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import json
from typing import Dict, List, Optional

class ExcelReportGenerator:
    """
    Automated Excel Report Generator
    Generates daily/weekly reports with formatting, charts, and multiple sheets
    """
    
    def __init__(self, config_file: str = "report_config.json"):
        self.config_file = config_file
        self.config = self._load_config()
        self.workbook = None
        
    def _load_config(self) -> Dict:
        """Load configuration from JSON file"""
        default_config = {
            "company_name": "Your Company Name",
            "report_title": "Automated Report",
            "output_directory": "reports",
            "date_format": "%Y-%m-%d",
            "colors": {
                "header": "366092",
                "subheader": "5B9BD5",
                "accent": "70AD47"
            }
        }
        
        if os.path.exists(self.config_file):
            with open(self.config_file, 'r') as f:
                user_config = json.load(f)
                default_config.update(user_config)
        else:
            # Create default config file
            with open(self.config_file, 'w') as f:
                json.dump(default_config, f, indent=4)
                
        return default_config
    
    def generate_sample_data(self, report_type: str = "daily") -> Dict[str, pd.DataFrame]:
        """Generate sample data for demonstration"""
        if report_type == "daily":
            # Daily sales data
            dates = pd.date_range(start=datetime.now() - timedelta(days=30), 
                                end=datetime.now(), freq='D')
            
            sales_data = pd.DataFrame({
                'Date': dates,
                'Product_A': np.random.randint(50, 200, len(dates)),
                'Product_B': np.random.randint(30, 150, len(dates)),
                'Product_C': np.random.randint(20, 100, len(dates)),
                'Total_Revenue': np.random.randint(1000, 5000, len(dates))
            })
            
            # Customer data
            customer_data = pd.DataFrame({
                'Customer_ID': [f'CUST_{i:04d}' for i in range(1, 101)],
                'Customer_Name': [f'Customer {i}' for i in range(1, 101)],
                'Total_Orders': np.random.randint(1, 20, 100),
                'Total_Spent': np.random.randint(100, 10000, 100),
                'Category': np.random.choice(['Premium', 'Standard', 'Basic'], 100)
            })
            
        else:  # weekly
            # Weekly performance data
            weeks = pd.date_range(start=datetime.now() - timedelta(weeks=12), 
                                end=datetime.now(), freq='W')
            
            sales_data = pd.DataFrame({
                'Week': weeks,
                'Sales_Target': np.random.randint(10000, 20000, len(weeks)),
                'Actual_Sales': np.random.randint(8000, 22000, len(weeks)),
                'New_Customers': np.random.randint(20, 100, len(weeks)),
                'Customer_Retention': np.random.uniform(0.7, 0.95, len(weeks))
            })
            
            # Department performance
            departments = ['Sales', 'Marketing', 'Support', 'Operations']
            customer_data = pd.DataFrame({
                'Department': departments,
                'Budget': [50000, 30000, 25000, 40000],
                'Actual_Spend': np.random.randint(20000, 55000, len(departments)),
                'Performance_Score': np.random.uniform(0.6, 1.0, len(departments)),
                'Team_Size': [10, 8, 12, 15]
            })
        
        return {
            'sales_data': sales_data,
            'customer_data': customer_data
        }
    
    def _apply_header_style(self, worksheet, row_num: int, col_range: str):
        """Apply header styling to worksheet"""
        header_fill = PatternFill(start_color=self.config['colors']['header'], 
                                end_color=self.config['colors']['header'], 
                                fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        
        for cell in worksheet[f'{col_range}{row_num}']:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    def _apply_data_formatting(self, worksheet, start_row: int, end_row: int, 
                             start_col: int, end_col: int):
        """Apply data formatting to worksheet"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Alternate row colors
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color='F2F2F2', 
                                          end_color='F2F2F2', 
                                          fill_type='solid')
    
    def _add_chart(self, worksheet, data_range: str, chart_type: str = 'bar', 
                  title: str = '', position: str = 'E2'):
        """Add chart to worksheet"""
        if chart_type == 'bar':
            chart = BarChart()
        else:
            chart = LineChart()
        
        chart.title = title
        chart.style = 13
        
        data = Reference(worksheet, range_string=data_range)
        chart.add_data(data, titles_from_data=True)
        
        worksheet.add_chart(chart, position)
    
    def create_summary_sheet(self, data: Dict[str, pd.DataFrame], 
                           report_type: str) -> None:
        """Create summary sheet with key metrics"""
        ws = self.workbook.create_sheet("Summary")
        
        # Title
        ws['A1'] = f"{self.config['company_name']} - {report_type.title()} Report Summary"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
        
        # Date
        ws['A2'] = f"Generated: {datetime.now().strftime(self.config['date_format'])}"
        ws['A2'].font = Font(italic=True)
        
        # Key metrics
        sales_data = data['sales_data']
        customer_data = data['customer_data']
        
        metrics_start_row = 4
        ws[f'A{metrics_start_row}'] = "Key Metrics"
        ws[f'A{metrics_start_row}'].font = Font(bold=True, size=14)
        
        if report_type == "daily":
            total_revenue = sales_data['Total_Revenue'].sum()
            avg_daily_revenue = sales_data['Total_Revenue'].mean()
            top_product = sales_data[['Product_A', 'Product_B', 'Product_C']].sum().idxmax()
            total_customers = len(customer_data)
            
            metrics = [
                ("Total Revenue", f"${total_revenue:,.2f}"),
                ("Average Daily Revenue", f"${avg_daily_revenue:,.2f}"),
                ("Top Product", top_product),
                ("Total Customers", total_customers),
                ("Premium Customers", len(customer_data[customer_data['Category'] == 'Premium']))
            ]
        else:  # weekly
            total_sales = sales_data['Actual_Sales'].sum()
            avg_performance = sales_data['Customer_Retention'].mean()
            total_new_customers = sales_data['New_Customers'].sum()
            
            metrics = [
                ("Total Sales", f"${total_sales:,.2f}"),
                ("Average Retention Rate", f"{avg_performance:.1%}"),
                ("Total New Customers", total_new_customers),
                ("Best Department", customer_data.loc[customer_data['Performance_Score'].idxmax(), 'Department'])
            ]
        
        for i, (metric, value) in enumerate(metrics, start=metrics_start_row + 1):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Apply formatting
        for row in ws.iter_rows(min_row=metrics_start_row + 1, 
                              max_row=metrics_start_row + len(metrics), 
                              min_col=1, max_col=2):
            for cell in row:
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
    
    def create_data_sheet(self, data: pd.DataFrame, sheet_name: str) -> None:
        """Create data sheet with formatted table"""
        ws = self.workbook.create_sheet(sheet_name)
        
        # Add data to worksheet
        for r in dataframe_to_rows(data, index=False, header=True):
            ws.append(r)
        
        # Apply formatting
        self._apply_header_style(ws, 1, f'A:{"ABCDEFGHIJKLMNOPQRSTUVWXYZ"[len(data.columns)-1]}')
        self._apply_data_formatting(ws, 2, len(data) + 1, 1, len(data.columns))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def generate_report(self, report_type: str = "daily", 
                       data: Optional[Dict[str, pd.DataFrame]] = None) -> str:
        """Generate complete Excel report"""
        
        # Create output directory if it doesn't exist
        os.makedirs(self.config['output_directory'], exist_ok=True)
        
        # Use provided data or generate sample data
        if data is None:
            data = self.generate_sample_data(report_type)
        
        # Create workbook
        self.workbook = Workbook()
        
        # Remove default sheet
        self.workbook.remove(self.workbook.active)
        
        # Create summary sheet
        self.create_summary_sheet(data, report_type)
        
        # Create data sheets
        for sheet_name, df in data.items():
            formatted_name = sheet_name.replace('_', ' ').title()
            self.create_data_sheet(df, formatted_name)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{report_type}_report_{timestamp}.xlsx"
        filepath = os.path.join(self.config['output_directory'], filename)
        
        # Save workbook
        self.workbook.save(filepath)
        
        return filepath

def main():
    """Main function to demonstrate usage"""
    # Initialize report generator
    generator = ExcelReportGenerator()
    
    print("🚀 Automated Excel Report Generator")
    print("=====================================")
    
    # Generate daily report
    print("\n📊 Generating daily report...")
    daily_report = generator.generate_report("daily")
    print(f"✅ Daily report saved: {daily_report}")
    
    # Generate weekly report
    print("\n📈 Generating weekly report...")
    weekly_report = generator.generate_report("weekly")
    print(f"✅ Weekly report saved: {weekly_report}")
    
    print("\n🎉 Reports generated successfully!")
    print(f"📁 Check the '{generator.config['output_directory']}' folder for your reports")

if __name__ == "__main__":
    main()
